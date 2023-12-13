from urllib.request import urlopen,Request
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font

# scrape the website below to retrieve the top 5 countries with the highest GDPs. Calculate the GDP per capita
# by dividing the GDP by the population. You can perform the calculation in Python natively or insert the code
# in excel that will perform the calculation in Excel by each row. DO NOT scrape the GDP per capita from the
# webpage, make sure you use your own calculation.

# FOR YOUR REFERENCE - https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html
# this link shows you the different number formats you can apply to a column using openpyxl


### REMEMBER ##### - your output should match the excel file (GDP_Report.xlsx) including all formatting.


webpage = 'https://www.worldometers.info/gdp/gdp-by-country/'

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}
req = Request(webpage, headers=headers)
webpage = urlopen(req).read()
soup = BeautifulSoup(webpage, 'html.parser')

# Find the table containing the data
table = soup.find('table', {'id': 'example2'})
data = [("No.", "Country", "GDP", "Population", "GDP Per Capita")]

for row in table.find_all('tr')[1:6]:  # Skip the header row
    columns = row.find_all('td')
    country = columns[1].text.strip()
    # Remove dollar sign and trillion
    gdp_raw = columns[2].text.replace('$', '').replace(' trillion', '').replace(',', '')
    population = (columns[5].text.replace(',', ''))
    
    # Convert trillion to the actual value and calculate GDP per capita
    gdp = int(gdp_raw) * 10**3
    population = int(population) * 1
    gdp_per_capita = round((gdp / population),2)

    data.append((len(data), country, gdp, population, gdp_per_capita))


wb = xl.Workbook()
ws = wb.active
ws.title = "GDP By Country"

#headers
headers = data[0]
ws.append(headers)

for row_data in data[1:]:
    ws.append(row_data)
# Remove the last 3 zeros from the GDP column
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
    for cell in row:
        cell.value = cell.value / 1000

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 30

bold_font = Font(bold=True, color="000000", size=16)  
Header_font = Font(bold=True, color="000000", size=16)  
for cell in ws['1:1']:
    cell.font = Header_font

# Column formats
column_formats = {
    "A": '#,##0',
    "B": None,
    "C": '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)',
    "D": '#,##0',
    "E": '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
}

for column, num_format in column_formats.items():
    for cell in ws[column]:
        if num_format:
            cell.number_format = num_format

# Save the workbook
wb.save("GDP_Report.xlsx")






